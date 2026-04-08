import streamlit as st
import pandas as pd
import io
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# =========================
# Configuración de página
# =========================
st.set_page_config(page_title="Cotizador UX/UI", page_icon="🧮", layout="wide")

# =========================
# Estilos CSS Avanzados (Tema Claro)
# =========================
def inyectar_css():
    st.markdown("""
        <style>
        /* Importar fuente moderna y corporativa */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

        /* Variables globales (Light Theme base) */
        :root {
            --primary-color: #0E2B5C;      /* Azul fuerte corporativo */
            --secondary-color: #3B82F6;    /* Azul brillante */
            --accent-color: #10B981;       /* Verde acento */
            --bg-color: #F8FAFC;           /* Fondo general más suave que el blanco puro */
            --card-bg: #FFFFFF;            /* Fondo de tarjetas */
            --text-main: #1E293B;          /* Texto oscuro para legibilidad */
            --text-muted: #64748B;         /* Texto secundario */
            --border-color: #E2E8F0;       /* Bordes muy sutiles */
        }

        /* Estilo base de Streamlit */
        .stApp {
            background-color: var(--bg-color);
            font-family: 'Inter', sans-serif !important;
            color: var(--text-main);
        }

        h1, h2, h3, h4, h5, h6, .stMarkdown, .stText {
            font-family: 'Inter', sans-serif !important;
        }

        h1, h2, h3 {
            color: var(--primary-color) !important;
        }

        /* Botones */
        button[kind="primary"] {
            background: linear-gradient(135deg, var(--secondary-color), var(--primary-color)) !important;
            color: white !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            padding: 0.6rem 1.2rem !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 6px -1px rgba(59, 130, 246, 0.2), 0 2px 4px -1px rgba(59, 130, 246, 0.1) !important;
        }
        
        button[kind="primary"]:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 10px 15px -3px rgba(59, 130, 246, 0.3), 0 4px 6px -2px rgba(59, 130, 246, 0.15) !important;
            opacity: 0.95 !important;
        }

        button[kind="secondary"] {
            background: rgba(255, 255, 255, 0.5) !important;
            color: var(--text-main) !important;
            border: 1px solid var(--border-color) !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
            transition: all 0.3s ease !important;
        }
        
        button[kind="secondary"]:hover {
            border-color: var(--secondary-color) !important;
            color: var(--secondary-color) !important;
            background: rgba(59, 130, 246, 0.05) !important;
            transform: translateY(-1px) !important;
        }

        /* Inputs de textos, selectbox y fechas */
        .stTextInput input, .stTextArea textarea, .stDateInput input, .stSelectbox select, .stNumberInput input {
            border-radius: 6px !important;
            border: 1px solid var(--border-color) !important;
            transition: border-color 0.2s, box-shadow 0.2s !important;
            background-color: var(--card-bg) !important;
        }

        .stTextInput input:focus, .stTextArea textarea:focus, .stDateInput input:focus, .stSelectbox select:focus, .stNumberInput input:focus {
            border-color: var(--secondary-color) !important;
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2) !important;
            outline: none !important;
        }

        /* Expander Title */
        .streamlit-expanderHeader {
            font-weight: 600 !important;
            color: var(--primary-color) !important;
            font-size: 1.1rem !important;
            background-color: var(--card-bg) !important;
            border-radius: 8px !important;
        }

        /* Tarjeta de Métricas custom */
        .metric-container {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            background-color: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 1.8rem;
            margin: 0.5rem 0;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease;
        }
        
        .metric-container:hover {
            transform: translateY(-4px);
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.08);
            border-color: #CBD5E1;
        }

        .metric-title {
            font-size: 0.95rem;
            color: var(--text-muted);
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            margin-bottom: 0.6rem;
            text-align: center;
        }

        .metric-value {
            font-size: 2rem;
            font-weight: 800;
            line-height: 1.2;
        }

        /* Colores semánticos sutiles pero claros */
        .val-22 { color: #3B82F6; }  /* Azul */
        .val-23 { color: #10B981; }  /* Verde */
        .val-25 { color: #F59E0B; }  /* Naranja */
        .val-30 { color: #EF4444; }  /* Rojo */

        /* Resaltar cabecera / Banner */
        .hero-banner {
            background: linear-gradient(120deg, var(--card-bg) 0%, #E0F2FE 100%);
            padding: 2.5rem;
            border-radius: 16px;
            margin-bottom: 2rem;
            border-left: 8px solid var(--secondary-color);
            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
        }
        .hero-banner h1 {
            color: var(--primary-color) !important;
            margin-top: 0 !important;
            font-size: 2.4rem;
            font-weight: 800 !important;
            margin-bottom: 0.5rem;
        }
        .hero-banner p {
            color: var(--text-muted);
            font-size: 1.15rem;
            font-weight: 400;
            margin-bottom: 0;
        }
        </style>
    """, unsafe_allow_html=True)

inyectar_css()

# =========================
# Cabecera Visual (Hero)
# =========================
st.markdown("""
<div class="hero-banner">
    <h1>🧮 Cotizador de Servicios UX/UI</h1>
    <p>Cálculo estructurado con márgenes de contribución para la planeación de proyectos de diseño.</p>
</div>
""", unsafe_allow_html=True)


# =========================
# Catálogo Estructurado
# Índices: [0]=22%, [1]=23%, [2]=25%, [3]=30%
# =========================
CATALOGO = {
    "DISEÑADOR UX/UI JR": {
        "Full": [126156, 127190, 129258, 134428],
        "Medio Tiempo": [75693, 76314, 77555, 80657]
    },
    "DISEÑADOR UX/UI MID": {
        "Full" : [126463, 127500, 129573, 134756],
        "Medio Tiempo": [75878, 76500, 77744, 80854]
    },
    "DISEÑADOR UX/UI SR": {
        "Full":[127500, 128545, 130635, 135861],
        "Medio Tiempo": [76500, 77127, 78381, 81516]
    },
    "PRODUCT DESIGNER": {
        "Full": [132283, 133367, 135536, 140957],
        "Medio Tiempo": [79370, 80020, 81322, 84574]
    },
    "SERVICE DESIGNER": {
        "Full": [147711, 148921, 151343, 157397],
        "Medio Tiempo": [88626, 89353, 90806, 94438]
    },
    "CUSTOMER SUCCESS": {
        "Full": [166777, 168144, 170878, 177713],
        "Medio Tiempo": [100066, 100886, 102527, 106628],
        "Medio Tiempo 30%": [50033, 50443, 51263, 53314]
    }
}

MONEDEROS = {
    "Tiendas Neto" : {
        "Monto": [200,300,400,500],
        "Monto con fee" : [200*1.05,300*1.05,400*1.05,500*1.05]
    },
    "Externo" : {
        "Monto": [200,300,400,500],
        "Monto con fee" : [200*1.15,300*1.15,400*1.15,500*1.15]
    }

}

# =========================
# Estado inicial (Session State)
# =========================
if "items_df" not in st.session_state:
    st.session_state.items_df = pd.DataFrame(
        columns=[
            "Rol", "Cant", "Meses", 
            "Precio 22%", "Precio 23%", "Precio 25%", "Precio 30%",
            "Subtotal 22%", "Subtotal 23%", "Subtotal 25%", "Subtotal 30%"
        ]
    )

if "datos" not in st.session_state:
    st.session_state.datos = {
        "Fecha de Cotizacion": date.today(),
        "Nombre del Cliente": "",
        "Proyecto": "",
        "Descripcion": "",
        "Tipo de Cliente": "Interno",
        "Contacto del Cliente": "",
        "Correo del Cliente":"",
        "Fecha de Inicio": date.today(),
        "Fecha de Fin": date.today(),
        "Entregables": "",
        "Antecedentes": "",
        "Presupuesto Cliente": "",
        "Target": "",
        "Objetivos Especificos": "",
        "Duracion Maxima": "",
        "Observaciones": ""
    }

if "monederos_list" not in st.session_state:
    st.session_state.monederos_list = []  # lista de dicts: {tipo, monto, monto_fee, personas}

def recalcular(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    # Asegurar tipos numéricos
    for col in ["Cant", "Meses", "Precio 22%", "Precio 23%", "Precio 25%", "Precio 30%"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # Recalcular totales: Precio * Cantidad * Meses
    for m in ["22%", "23%", "25%", "30%"]:
        df[f"Subtotal {m}"] = (df[f"Precio {m}"] * df["Cant"] * df["Meses"]).round(2)
    return df

# =========================
# 1) Datos generales
# =========================
st.markdown("### 📋 1. Información del Proyecto")

with st.expander("📝 Datos Generales del Cliente y Proyecto", expanded=True):
    col_fecha, col_empty = st.columns([1, 4])
    with col_fecha:
        fecha = st.date_input("📅 Fecha de cotización", value=st.session_state.datos["Fecha de Cotizacion"])
    
    st.markdown("<hr style='margin-top:0.5rem; margin-bottom:1.5rem; opacity:0.3;'>", unsafe_allow_html=True)
        
    col1, col2 = st.columns([1, 1], gap="large")
    with col1:
        cliente = st.text_input("🏢 Nombre del Cliente", value=st.session_state.datos["Nombre del Cliente"], placeholder="Ej. UPAX S.A. de C.V.")
        contacto = st.text_input("📞 Teléfono de contacto", value = st.session_state.datos["Contacto del Cliente"], placeholder = "+00 0000 0000")
        fecha_inicio = st.date_input("🚀 Fecha de inicio", value = st.session_state.datos["Fecha de Inicio"])
        presupuesto = st.text_input("💰 Presupuesto del cliente", value=st.session_state.datos["Presupuesto Cliente"], placeholder="Ej. $100,000 MXN")
        
    with col2:
        tipo_cliente = st.selectbox("🏷️ Tipo de Cliente", options=["Interno", "Externo"], index=0 if st.session_state.datos["Tipo de Cliente"] == "Interno" else 1)
        correo = st.text_input("✉️ Correo electrónico", value = st.session_state.datos["Correo del Cliente"], placeholder = "cliente@email.com")
        fecha_fin = st.date_input("🏁 Fecha de finalización", value = st.session_state.datos["Fecha de Fin"])
        target = st.text_input("🎯 Público Objetivo / Target", value=st.session_state.datos["Target"])
        
    proyecto = st.text_input("📂 Nombre del Proyecto", value=st.session_state.datos["Proyecto"])
    
    # Text areas layout
    colTA1, colTA2 = st.columns(2, gap="large")
    with colTA1:
        descripcion = st.text_area("📝 Descripción y Objetivo", value=st.session_state.datos["Descripcion"], height=120)
        entregables = st.text_area("📦 Entregables del proyecto", value=st.session_state.datos["Entregables"], placeholder="Ej. Prototipos, Journey Maps...", height=120)
        objetivos = st.text_area("🎯 Objetivos Específicos", value=st.session_state.datos["Objetivos Especificos"], placeholder="1. Reducir abandono...", height=120)

    with colTA2:
        antecedentes = st.text_area("📖 Antecedentes / Justificación", value=st.session_state.datos["Antecedentes"], placeholder="Contexto de negocio...", height=120)
        duracion_maxima = st.text_area("⏱️ Duración Máxima Estimada", value=st.session_state.datos["Duracion Maxima"], placeholder="Ej. El proyecto tendrá duración de 4 meses...", height=120)
        observaciones = st.text_area("🔍 Observaciones Adicionales", value=st.session_state.datos["Observaciones"], placeholder="Notas internas...", height=120)

# Sincronizar datos
st.session_state.datos.update({
    "Fecha de Cotizacion": fecha, "Nombre del Cliente": cliente, "Proyecto": proyecto, "Descripcion": descripcion,
    "Tipo de Cliente": tipo_cliente, "Contacto del Cliente": contacto, "Correo del Cliente": correo,
    "Fecha de Inicio": fecha_inicio, "Fecha de Fin": fecha_fin, "Entregables": entregables,
    "Antecedentes": antecedentes, "Presupuesto Cliente": presupuesto, "Target": target, 
    "Objetivos Especificos": objetivos, "Duracion Maxima": duracion_maxima, "Observaciones": observaciones
})

st.markdown("<br>", unsafe_allow_html=True)

# =========================
# 2) Agregar recursos
# =========================
st.markdown("### 👥 2. Asignación de Recursos")

#st.markdown("<div style='background-color:#FFFFFF; padding:2rem; border-radius:12px; border: 1px solid #E2E8F0; margin-bottom: 2rem;'>", unsafe_allow_html=True)
colA, colB, colC = st.columns([1.5, 1, 1], gap="medium")

with colA:
    rol_sel = st.selectbox("Perfil del Especialista", options=list(CATALOGO.keys()))
    opciones_dedicacion = list(CATALOGO[rol_sel].keys())
    tiempo_sel = st.radio("Dedicación", options=opciones_dedicacion, horizontal=True)

# Extraer los 4 precios del catálogo
precios = CATALOGO[rol_sel][tiempo_sel]

with colB:
    cantidad = st.number_input("Cantidad de personas", min_value=1, value=1)
    st.info(f"Tarifa Minima (22%): **${precios[0]:,}**", icon="ℹ️")
    
with colC:
    meses = st.number_input("Meses", min_value=0.1, value=1.0, step=0.5)
    st.error(f"Tarifa Máxima (30%): **${precios[3]:,}**", icon="📈")

colBtnA, _ = st.columns([1, 2])
with colBtnA:
    if st.button("➕ Agregar recurso al presupuesto", type="primary", use_container_width=True):
        factor = cantidad * meses
        nuevo = pd.DataFrame([{
            "Rol": f"{rol_sel} ({tiempo_sel})",
            "Cant": int(cantidad),
            "Meses": float(meses),
            "Precio 22%": precios[0], "Precio 23%": precios[1], "Precio 25%": precios[2], "Precio 30%": precios[3],
            "Subtotal 22%": round(precios[0] * factor, 2),
            "Subtotal 23%": round(precios[1] * factor, 2),
            "Subtotal 25%": round(precios[2] * factor, 2),
            "Subtotal 30%": round(precios[3] * factor, 2)
        }])
        st.session_state.items_df = pd.concat([st.session_state.items_df, nuevo], ignore_index=True)
        st.rerun()

st.markdown("</div>", unsafe_allow_html=True)

st.markdown("### 👛 3. Monederos")

incluir_monederos = st.toggle("Incluir Monederos en la cotización", value=False, key="toggle_monederos")

if incluir_monederos:
    colM1, colM2, colM3 = st.columns([1.5, 1, 1], gap="medium")

    with colM1:
        tipo_monedero = st.selectbox("🏦 Tipo de Monedero", options=list(MONEDEROS.keys()), key="sel_tipo_monedero")
        montos_disponibles = MONEDEROS[tipo_monedero]["Monto"]
        montos_con_fee = MONEDEROS[tipo_monedero]["Monto con fee"]
        fee_pct = "5%" if tipo_monedero == "Tiendas Neto" else "15%"

    with colM2:
        monto_idx = st.selectbox(
            "💵 Monto por Persona",
            options=range(len(montos_disponibles)),
            format_func=lambda i: f"${montos_disponibles[i]:,.0f}",
            key="sel_monto_monedero"
        )

    with colM3:
        personas_monedero = st.number_input("👤 Número de Personas", min_value=1, value=1, key="num_personas_monedero")
        costo_total_monedero = montos_con_fee[monto_idx] * personas_monedero
        st.success(f"Costo total **${costo_total_monedero:,.2f}**", icon="🧾")

    colBtnM, _ = st.columns([1, 2])
    with colBtnM:
        if st.button("➕ Agregar monedero al presupuesto", type="primary", use_container_width=True, key="btn_add_monedero"):
            st.session_state.monederos_list.append({
                "Tipo": tipo_monedero,
                "Monto Base": montos_disponibles[monto_idx],
                "Fee": fee_pct,
                "Monto c/Fee": round(montos_con_fee[monto_idx], 2),
                "Personas": int(personas_monedero),
                "Total c/Fee": round(costo_total_monedero, 2)
            })
            st.rerun()

    # Mostrar tabla de monederos agregados
    if st.session_state.monederos_list:
        st.markdown("<p style='color: var(--text-muted); font-size:0.9rem; margin-top:1rem;'><em>Monederos agregados a la cotización:</em></p>", unsafe_allow_html=True)
        df_monederos = pd.DataFrame(st.session_state.monederos_list)
        st.dataframe(df_monederos, use_container_width=True, hide_index=True)

        colLimpiaM, _ = st.columns([1, 4])
        with colLimpiaM:
            if st.button("🗑️ Limpiar monederos", use_container_width=True, key="btn_limpiar_monederos"):
                st.session_state.monederos_list = []
                st.rerun()
    else:
        st.info("No hay monederos agregados. Selecciona el tipo, monto y número de personas y presiona el botón.", icon="👛")
else:
    # Si el toggle está apagado, limpiar la lista para que no afecte los totales
    if st.session_state.monederos_list:
        st.session_state.monederos_list = []

# Calcular costo total de monederos (se usará en el resumen)
total_monederos_fee = sum(m["Total c/Fee"] for m in st.session_state.monederos_list)


# =========================
# 3) Detalle y Totales
# =========================
st.markdown("### 📊 4. Resumen y Previsualización")


# Tabla interactiva
st.markdown("<p style='color: var(--text-muted); font-size: 0.95rem;'><em>Puedes editar directamente las Cantidades y Meses en la siguiente tabla.</em></p>", unsafe_allow_html=True)
edited_df = st.data_editor(
    st.session_state.items_df,
    num_rows="dynamic",
    use_container_width=True,
    key="editor_tabla"
)

if not edited_df.equals(st.session_state.items_df):
    st.session_state.items_df = recalcular(edited_df)
    st.rerun()

# Cálculos finales
totales = st.session_state.items_df[["Subtotal 22%", "Subtotal 23%", "Subtotal 25%", "Subtotal 30%"]].sum()

# Totales incluyendo monederos con fee
total_22_con_mon = totales['Subtotal 22%'] + total_monederos_fee
total_23_con_mon = totales['Subtotal 23%'] + total_monederos_fee
total_25_con_mon = totales['Subtotal 25%'] + total_monederos_fee
total_30_con_mon = totales['Subtotal 30%'] + total_monederos_fee

st.markdown("<br>", unsafe_allow_html=True)
st.markdown("#### Totales por Margen de Contribución")

# Sub-etiqueta de monedero
mon_label = f" <span style='font-size:0.8rem; color:#64748B;'>(incl. monedero ${total_monederos_fee:,.2f})</span>" if total_monederos_fee > 0 else ""

# Implementación de HTML custom para las tarjetas de métricas
html_cards = f"""
<div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 1.5rem; margin-bottom: 2rem;">
    <div class="metric-container">
        <div class="metric-title">Mínimo (22%) {mon_label}</div>
        <div class="metric-value val-22">${total_22_con_mon:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Recursos: ${totales['Subtotal 22%']:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Monedero: ${total_monederos_fee:,.2f}</div>
    </div>
    <div class="metric-container">
        <div class="metric-title">Base (23%) {mon_label}</div>
        <div class="metric-value val-23">${total_23_con_mon:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Recursos: ${totales['Subtotal 23%']:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Monedero: ${total_monederos_fee:,.2f}</div>
    </div>
    <div class="metric-container">
        <div class="metric-title">Óptimo (25%) {mon_label}</div>
        <div class="metric-value val-25">${total_25_con_mon:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Recursos: ${totales['Subtotal 25%']:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Monedero: ${total_monederos_fee:,.2f}</div>
      </div>
    <div class="metric-container">
        <div class="metric-title">Máximo (30%) {mon_label}</div>
        <div class="metric-value val-30">${total_30_con_mon:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Recursos: ${totales['Subtotal 30%']:,.2f}</div>
        <div style="font-size:0.8rem; color:#94A3B8; margin-top:0.3rem;">Monedero: ${total_monederos_fee:,.2f}</div>
    </div>
</div>
"""
st.markdown(html_cards, unsafe_allow_html=True)

st.warning(f"**⚠️ Regla de Negocio:** El total final (recursos + monederos) no debe ser menor (\${total_22_con_mon:,.2f}) (22%) ni mayor (\${total_30_con_mon:,.2f}) (30%)", icon="🚨")

st.markdown("<br>", unsafe_allow_html=True)

colLimpia, _ = st.columns([1, 4])
with colLimpia:
    if st.button("🗑️ Limpiar todos los recursos", use_container_width=True):
        st.session_state.items_df = st.session_state.items_df.iloc[0:0]
        st.rerun()

st.divider()

# =========================
# 4) Exportar a Excel
# =========================



def generar_excel(datos, df, monederos_list=None):
    output = io.BytesIO()
    if monederos_list is None:
        monederos_list = []
    
    # 🎨 Definición de Estilos (Colores Corporativos)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0E2B5C", end_color="0E2B5C", fill_type="solid")
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    wrap_aligned_text = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"), 
        right=Side(style='thin', color="E2E8F0"), 
        top=Side(style='thin', color="E2E8F0"), 
        bottom=Side(style='thin', color="E2E8F0")
    )
    accent_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    monedero_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    totales_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ==========================================
        # --- Hoja 1: Datos Generales ---
        # ==========================================
        datos_formateados = [{"CAMPO": k.upper(), "VALOR": str(v)} for k, v in datos.items() if v]
        pd.DataFrame(datos_formateados).to_excel(writer, sheet_name="Datos Generales", index=False)
        
        ws1 = writer.sheets["Datos Generales"]
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 60
        
        for cell in ws1["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned_text
            cell.border = thin_border
            
        for row in ws1.iter_rows(min_row=2, max_col=2, max_row=ws1.max_row):
            for cell in row:
                cell.alignment = wrap_aligned_text
                cell.border = thin_border
                if cell.column == 1:
                    cell.font = Font(bold=True, color="64748B")

        # ==========================================
        # --- Hoja 2: Cotización ---
        # ==========================================
        df.to_excel(writer, sheet_name="Cotización", index=False)
        ws = writer.sheets["Cotización"]
        
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['A'].width = 30
        
        for cell in ws["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned_text
            cell.border = thin_border
            
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if cell.column >= 4:
                    cell.number_format = '"$"#,##0.00'
        
        # ==========================================
        # --- Sección de Monederos en Excel ---
        # ==========================================
        total_monederos_excel = 0
        if monederos_list:
            row_mon_titulo = ws.max_row + 2
            
            # Título de sección
            t_cell = ws.cell(row=row_mon_titulo, column=1, value="👛 MONEDEROS")
            t_cell.font = Font(bold=True, color="0E2B5C", size=11)
            t_cell.fill = accent_fill
            ws.merge_cells(start_row=row_mon_titulo, start_column=1, end_row=row_mon_titulo, end_column=6)

            # Cabecera de monederos
            mon_headers = ["Tipo", "Monto Base", "# de Monederos", "Total"]
            row_mon_header = row_mon_titulo + 1
            for ci, h in enumerate(mon_headers, start=1):
                c = ws.cell(row=row_mon_header, column=ci, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                c.alignment = center_aligned_text
                c.border = thin_border

            # Filas de monederos
            for ri, mon in enumerate(monederos_list, start=row_mon_header + 1):
                vals = [mon["Tipo"], mon["Monto Base"], mon["Personas"], mon["Total c/Fee"]]
                for ci, v in enumerate(vals, start=1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.border = thin_border
                    c.fill = monedero_fill
                    c.alignment = Alignment(vertical="center", horizontal="center")
                    if ci in (2, 4, 6):  # columnas monetarias
                        c.number_format = '"$"#,##0.00'
                total_monederos_excel += mon["Total c/Fee"]

            # Fila de total de monederos
            row_mon_total = row_mon_header + len(monederos_list) + 1
            lbl = ws.cell(row=row_mon_total, column=3, value="TOTAL MONEDEROS")
            lbl.font = Font(bold=True, color="0E2B5C")
            lbl.alignment = Alignment(horizontal="right", vertical="center")
            lbl.fill = accent_fill
            lbl.border = thin_border
            val_mon = ws.cell(row=row_mon_total, column=4, value=total_monederos_excel)
            val_mon.number_format = '"$"#,##0.00'
            val_mon.font = Font(bold=True, size=11, color="0E2B5C")
            val_mon.fill = accent_fill
            val_mon.border = thin_border
            val_mon.alignment = center_aligned_text

        # ==========================================
        # --- Totales de Recursos + Monederos ---
        # ==========================================
        row_titulos = ws.max_row + 2
        row_valores = row_titulos + 1
        
        titulo_cell = ws.cell(row=row_titulos, column=1, value="RESUMEN DE TOTALES (Recursos + Monederos)")
        titulo_cell.font = Font(bold=True, color="0E2B5C", size=12)
        ws.merge_cells(start_row=row_titulos, start_column=1, end_row=row_titulos, end_column=7)
        
        columnas_sumar = ["Subtotal 22%", "Subtotal 23%", "Subtotal 25%", "Subtotal 30%"]
        totales_sum = df[columnas_sumar].sum()
        
        for i, col_name in enumerate(columnas_sumar, start=8):
            c_header = ws.cell(row=row_titulos, column=i, value=f"Total {col_name.split()[-1]}")
            c_header.font = Font(bold=True, color="64748B")
            c_header.fill = totales_fill
            c_header.alignment = center_aligned_text
            c_header.border = thin_border
            
            valor_final = totales_sum[col_name] + total_monederos_excel
            c_val = ws.cell(row=row_valores, column=i, value=valor_final)
            c_val.number_format = '"$"#,##0.00'
            c_val.font = Font(bold=True, size=12, color="1E293B")
            c_val.border = thin_border
            c_val.alignment = center_aligned_text
            c_val.fill = totales_fill
            
        # Mensaje de Advertencia
        t22 = totales_sum['Subtotal 22%'] + total_monederos_excel
        t30 = totales_sum['Subtotal 30%'] + total_monederos_excel
        msg = f"⚠️ ADVERTENCIA: El total final (recursos + monederos) no debe ser menor (${t22:,.2f}) (22%) ni mayor (${t30:,.2f}) (30%)"
        msg_cell = ws.cell(row=row_titulos + 3, column=1, value=msg)
        msg_cell.font = Font(bold=True, color="EF4444")
        ws.merge_cells(start_row=row_titulos + 3, start_column=1, end_row=row_titulos + 3, end_column=11)
            
    return output.getvalue()

def enviar_correo(destinatario, asunto, cuerpo, archivo_bytes, nombre_archivo):
    remitente = st.secrets["email"]["cotizacion"]
    password = st.secrets["email"]["cotizacion_pass"]
    #remitente = "calculadora.cotizacion.uix@gmail.com"
    #password = "xstj flnb otsf vmfm"
    
    msg = MIMEMultipart()
    msg['From'] = remitente
    msg['To'] = destinatario
    msg['Subject'] = asunto
    msg.attach(MIMEText(cuerpo, 'plain'))

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(archivo_bytes)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f"attachment; filename= {nombre_archivo}")
    msg.attach(part)

    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remitente, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception:
        return False

def procesar_descarga_silenciosa(datos, xlsx_data, file_name):
    lista_correos = [st.secrets["email"]["correo_1"], st.secrets["email"]["correo_2"]]
    #lista_correos = ["oswaldoraulsanchez@gmail.com"]
    asunto = f"Cotización Proyecto: {datos['Proyecto']}"
    cuerpo = f"Hola,\n\nAdjunto se envía la cotización para el proyecto {datos['Proyecto']} del cliente {datos['Nombre del Cliente']}.\n\nSaludos."
    
    for destinatario in lista_correos:
        enviar_correo(destinatario, asunto, cuerpo, xlsx_data, file_name)

st.markdown("### 📥 4. Generar Documentación")
st.markdown("<p style='color: var(--text-muted); font-size: 0.95rem;'>Completa la información del proyecto y agrega recursos para habilitar la descarga en Excel.</p>", unsafe_allow_html=True)

if not st.session_state.items_df.empty and (st.session_state.datos["Nombre del Cliente"] and st.session_state.datos["Fecha de Cotizacion"] and st.session_state.datos["Proyecto"] and st.session_state.datos["Descripcion"] and st.session_state.datos["Tipo de Cliente"] and st.session_state.datos["Contacto del Cliente"]):
    xlsx_data = generar_excel(st.session_state.datos, st.session_state.items_df, st.session_state.monederos_list)
    file_name = f"Cotizacion_{st.session_state.datos['Nombre del Cliente']}_{st.session_state.datos['Fecha de Cotizacion']}.xlsx".replace(" ", "_")

    colDescarga, _ = st.columns([1, 2])
    with colDescarga:
        st.download_button(
            label="⬇️ Descargar Reporte en Excel",
            data=xlsx_data,
            file_name=file_name,
            use_container_width=True,
            type="primary",
            on_click=procesar_descarga_silenciosa,
            args=(st.session_state.datos, xlsx_data, file_name)
        )
else:
    st.info("Para habilitar la descarga, asegúrate de haber capturado toda la información básica (Cliente, Proyecto, etc.) y tener al menos un recurso en la tabla.", icon="💡")

